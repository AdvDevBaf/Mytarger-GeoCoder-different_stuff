import requests
import hashlib
import json
from requests_toolbelt.multipart.encoder import MultipartEncoder
from tkinter import *
from tkinter.filedialog import askopenfilename
import openpyxl
import ctypes
import time


def refill_table(block):
    wb = openpyxl.load_workbook(str('C:\\Users\\AMasanov\\Desktop\\online_cinema.xlsx'))
    sheet = wb.worksheets[0]
    for i in range(0, len(block)):
        print(block[i])
        sheet.cell(row=i + 1, column=3).value = str(block[i])
    wb.save(str('C:\\Users\\AMasanov\\Desktop\\online_cinema.xlsx'))
    message = 'Готово!'
    ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
    print('ok')
    return {}


def get_columns_data():
    # Получим столбцы из эксель файла
    try:
        wb = openpyxl.load_workbook(str('C:\\Users\\AMasanov\\Desktop\\online_cinema.xlsx'))
    except FileNotFoundError:
        message = 'Укажите путь к таблице!'
        ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
    sheet = wb.worksheets[0]
    address = []
    max_row = sheet.max_row
    for i in range(1, sheet.max_row):
        if sheet.cell(row=i, column=1).value is None:
            max_row = i - 1
            print(max_row)
            break
        else:
            max_row = sheet.max_row
            print(max_row)

    for i in range(1, max_row + 1):
        address.append(sheet.cell(row=i, column=1).value)

    return address


def get_blocked(address):
    blocked = []
    for adr in address:
        try:
            response = requests.get(str('http://')+str(adr))
            print(response.status_code)
            if 'Access to information resources is restricted on the basis of the Federal Law' in response.text:
                print('denied')
                blocked.append('заблокирован или недоступен')
                time.sleep(5)
            else:
                blocked.append('доступен')
        except:
            print('not http')
            blocked.append('заблокирован или недоступен')
    return blocked


address = get_columns_data()
blocks = get_blocked(address)
s = refill_table(blocks)
