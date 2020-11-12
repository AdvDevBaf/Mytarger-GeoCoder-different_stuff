import json
from tkinter import *
from tkinter.filedialog import askopenfilename
import requests
import openpyxl
import ctypes
import time


class Gui(Toplevel):
    def __init__(self, parent, title="Create ads & campaigns"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def search_folder_for_new_excel_file(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def get_columns_data(self):
        from datetime import datetime
        # Get cells from excel file
        wb = openpyxl.load_workbook(str(self.text_1.get()))
        sheet = wb.worksheets[0]

        # Create lists for campaigns
        name = []
        company_day_limit = []
        company_all_limit = []
        start_time = []
        stop_time = []

        cpm = []
        cpc = []
        sex = []
        age_from = []
        age_to = []
        geo = []
        link_template = []
        link_url = []
        pixels = []

        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value is None:
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(2, max_row + 1):
            # Fill campaign lists
            name.append(sheet.cell(row=i, column=1).value)
            sex.append(sheet.cell(row=i, column=2).value)
            age_from.append(sheet.cell(row=i, column=3).value)
            age_to.append(sheet.cell(row=i, column=4).value)
            geo.append(str(sheet.cell(row=i, column=5).value))
            company_day_limit.append(str(sheet.cell(row=i, column=8).value))
            company_all_limit.append(str(sheet.cell(row=i, column=9).value))
            cpm.append(str(sheet.cell(row=i, column=10).value).replace(',', '.'))
            cpc.append(str(sheet.cell(row=i, column=11).value).replace(',', '.'))
            link_template.append(str(sheet.cell(row=i, column=12).value).replace('\\', '/'))
            link_url.append(str(sheet.cell(row=i, column=13).value).replace('\\', '/'))
            pixels.append(str(sheet.cell(row=i, column=14).value).replace('\\', '/'))

            start = datetime.strptime(str(sheet.cell(row=i, column=6).value)[:10].replace('-', '.'), '%Y.%m.%d')
            stop = datetime.strptime(str(sheet.cell(row=i, column=7).value)[:10].replace('-', '.'), '%Y.%m.%d')
            start_time.append(str(start.strftime('%d.%m.%Y')))
            stop_time.append(str(stop.strftime('%d.%m.%Y')))

        print(start_time)
        print(stop_time)
        print(cpm)
        return name, company_day_limit, company_all_limit, start_time, stop_time,\
               cpm, cpc, link_url, pixels, link_template, sex, age_from, age_to, geo

    def get_columns_data_cpc(self):
        from datetime import datetime
        # Get cells from excel file
        wb = openpyxl.load_workbook(str(self.text_1.get()))
        sheet = wb.worksheets[0]

        # Create lists for campaigns
        name = []
        company_day_limit = []
        company_all_limit = []
        start_time = []
        stop_time = []

        cpm = []
        cpc = []
        sex = []
        age_from = []
        age_to = []
        geo = []
        link_template_vorota = []
        link_url_vorota = []
        link_template_goal = []
        link_url_goal = []
        link_template_mobil1 = []
        link_url_mobil1 = []
        pixels = []

        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value is None:
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(2, max_row + 1):
            # Fill campaign lists
            name.append(sheet.cell(row=i, column=1).value)
            sex.append(sheet.cell(row=i, column=2).value)
            age_from.append(sheet.cell(row=i, column=3).value)
            age_to.append(sheet.cell(row=i, column=4).value)
            geo.append(str(sheet.cell(row=i, column=5).value))
            company_day_limit.append(str(sheet.cell(row=i, column=8).value))
            company_all_limit.append(str(sheet.cell(row=i, column=9).value))
            cpm.append(str(sheet.cell(row=i, column=10).value).replace(',', '.'))
            cpc.append(str(sheet.cell(row=i, column=11).value).replace(',', '.'))
            link_template_vorota.append(str(sheet.cell(row=i, column=12).value).replace('\\', '/'))
            link_url_vorota.append(str(sheet.cell(row=i, column=13).value).replace('\\', '/'))
            link_template_goal.append(str(sheet.cell(row=i, column=14).value).replace('\\', '/'))
            link_url_goal.append(str(sheet.cell(row=i, column=15).value).replace('\\', '/'))
            link_template_mobil1.append(str(sheet.cell(row=i, column=16).value).replace('\\', '/'))
            link_url_mobil1.append(str(sheet.cell(row=i, column=17).value).replace('\\', '/'))
            pixels.append(str(sheet.cell(row=i, column=18).value).replace('\\', '/'))

            start = datetime.strptime(str(sheet.cell(row=i, column=6).value)[:10].replace('-', '.'), '%Y.%m.%d')
            stop = datetime.strptime(str(sheet.cell(row=i, column=7).value)[:10].replace('-', '.'), '%Y.%m.%d')
            start_time.append(str(start.strftime('%d.%m.%Y')))
            stop_time.append(str(stop.strftime('%d.%m.%Y')))


        print(start_time)
        print(stop_time)
        print(cpm)
        return name, company_day_limit, company_all_limit, start_time, stop_time,\
               cpm, cpc, link_template_vorota, link_url_vorota, \
               link_template_goal, link_url_goal, link_template_mobil1, link_url_mobil1, pixels, \
               sex, age_from, age_to, geo

    def __refresh_token(self):
        pass  # Получить/Обновить токен

    def __get_geo(self, geo):
        address = []
        radius = []
        lng = []
        lat = []
        for element in geo:
            location = element.split(':')
            location = location[::-1]
            address.append(str(location[2]))
            radius.append(str(location[1])[:len(location[1])-1])
            coord = str(location[0]).split(',')
            lng.append(coord[1])
            lat.append(coord[0])
        print(address)
        print(radius)
        print(lng)
        print(lat)
        return address, radius, lng, lat

    def __create_campaigns(self, name, company_day_limit, company_all_limit, start_time, stop_time, cpm, cpc,
                           sex, age_from, age_to, address, radius, lng, lat, pixels, token):
        print(age_from[0])
        print(age_to[0])
        count = 0
        campaign_ids = []
        campaign_names = []
        __header = {'Host': 'target.my.com', 'Content-Type': 'application/json',
                    'Accept-Encoding': 'gzip,deflate,compress',
                    'Authorization': 'Bearer 7Db55KfCPP0vpYzyWna8mmfOwGwSmizWGgDECvNzT4OMSNIL1c3BqyWuJKyC5sr8ahL2oPDbGoRvYhRlOHblrcnnVQSjSB463vc2xTTU3l9Z8HckazKcMWJcpQFWgHPLESG3t5TsbJ0m1IhK5c2toMT3Zc8LODINPlGPDDgeNYkgwMLtV9hQGfZWkwCOFL6LCPiofY9bGKtNeHMtNH94cxAa5uHHtomY'}

        __sex = {
            "Мужской": "M",
            "Женский": "F",
            "Любой": "MF"
        }

        # 503 20290 20292 | 504 20291 20289
        # desktop_site_240x400_cpc_ttm_mailru
        for i in range(len(name)):
            if str(name[i]) not in campaign_names:
                campaign_names.append(str(name[i]))
                if str(cpm[i]) == 'None':
                    print('da')
                    company = {"name": str(name[i]), "price": str(cpc[i]), "budget_limit_day": str(company_day_limit[i]),
                               "budget_limit": str(company_all_limit[i]), "date_start": str(start_time[i]),
                               "date_end": str(stop_time[i]), "package": {"id": 503},
                               "targetings": {"regions": [188], "sex": __sex[sex[i]], "interests": [7367, 7390, 7392,
                                                                                                    7394, 8882,
                                                                                                    11213, 10268],
                                              "age": [x for x in range(age_from[i], age_to[i] + 1)],
                                              "pads": [{'id': 20290}],
                                              "local_geo": {"type": "now", "regions": [
                                                  {"lng": str(lng[i]),  # долгота
                                                   "lat": str(lat[i]),  # широта
                                                   "radius": str(radius[i]),
                                                   "label": str(address[i])}]
                                                                                                }
                                              }
                               }
                else:
                    print('net')
                    company = {"name": str(name[i]), "price": str(cpm[i]), "budget_limit_day": str(company_day_limit[i]),
                               "budget_limit": str(company_all_limit[i]), "date_start": str(start_time[i]),
                               "date_end": str(stop_time[i]), "package": {"id": 504},
                               "audit_pixels": {'impression': [str(pixels[i])]},
                               "targetings": {"regions": [188], "sex": __sex[sex[i]], "interests": [7367, 7390, 7392,
                                                                                                    7394, 8882,
                                                                                                    11213, 10268],
                                              "age": [x for x in range(age_from[i], age_to[i]+1)],
                                              "pads": [{"id": 20291}, {'id': 20289}],
                                              "local_geo": {"type": "now", "regions": [{"lng": str(lng[i]),  # долгота
                                                                                        "lat": str(lat[i]),  # широта
                                                                                        "radius": str(radius[i]),
                                                                                        "label": str(address[i])}]
                                                            }
                                              }
                               }


                response = requests.post('https://target.my.com/api/v1/campaigns.json', data=json.dumps(company),
                                         headers=__header)
                count += 1
                print(count)
                if count == 179:
                    import datetime
                    now = datetime.datetime.now()
                    a = (60 - now.minute)*60
                    print(a)
                    time.sleep(a)
                print(response.json())
                print(response.json()['id'])
                campaign_ids.append(response.json()['id'])
                print(campaign_ids)
            else:
                campaign_ids.append(campaign_names.index(str(name[i])))

        return campaign_ids

    def __create_template_ids(self, link_template):
        template_ids = []
        headers = {
            'Host': 'target.my.com',
            'Authorization': 'Bearer 7Db55KfCPP0vpYzyWna8mmfOwGwSmizWGgDECvNzT4OMSNIL1c3BqyWuJKyC5sr8ahL2oPDbGoRvYhRlOHblrcnnVQSjSB463vc2xTTU3l9Z8HckazKcMWJcpQFWgHPLESG3t5TsbJ0m1IhK5c2toMT3Zc8LODINPlGPDDgeNYkgwMLtV9hQGfZWkwCOFL6LCPiofY9bGKtNeHMtNH94cxAa5uHHtomY',
            'Accept-Encoding': 'gzip,deflate,compress',
            # 'Content-Type': 'multipart/form-data; boundary=----sample'
        }
        for i in range(len(link_template)):
            file = {
                'file': (str(link_template[i]),
                         open(str(link_template[i]), 'rb'))
            }
            response = requests.post('https://target.my.com/api/v2/content/html5.json', files=file,
                                     headers=headers)
            print(response.json())
            template_ids.append(str(response.json()['id']))
        return template_ids

    def __create_template_ids_cpc(self, link_template_vorota, link_template_goal, link_template_mobil1):
        template_ids = [[], [], []]
        headers = {
            'Host': 'target.my.com',
            'Authorization': 'Bearer 75kH6nlRjszveEBvR1jjH8667ykcluDt9BabULHPECGNwcyRKonvBPvBpAdkwD7tY2Ukocy0Qq6AdDxcKOTi6zrnjd9NwX3Igp8eh0eGw6AM1dopXtze6bk6Q22XyWzZ3H0XfulieOkRg0SPshLrNFdC60mIEd4P3C9VUoU2KgGrYBu3OKF1weUsEp8p9Bo4SQyQx1ofWMvTUzDv0q2IjczfB',
            'Accept-Encoding': 'gzip,deflate,compress',
            # 'Content-Type': 'multipart/form-data; boundary=----sample'
        }
        for i in range(len(link_template_vorota)):
            file_vorota = {
                'file': (str(link_template_vorota[i]),
                         open(str(link_template_vorota[i]), 'rb'))
            }
            response_vorota = requests.post('https://target.my.com/api/v2/content/html5.json', files=file_vorota,
                                     headers=headers)

            file_goal = {
                'file': (str(link_template_goal[i]),
                         open(str(link_template_goal[i]), 'rb'))
            }
            response_goal = requests.post('https://target.my.com/api/v2/content/html5.json', files=file_goal,
                                     headers=headers)

            file_mobil1 = {
                'file': (str(link_template_mobil1[i]),
                         open(str(link_template_mobil1[i]), 'rb'))
            }
            response_mobil1 = requests.post('https://target.my.com/api/v2/content/html5.json', files=file_mobil1,
                                          headers=headers)

            print(response_vorota.json())
            print(response_goal.json())
            print(response_mobil1.json())
            template_ids[0].append(str(response_vorota.json()['id']))
            template_ids[1].append(str(response_goal.json()['id']))
            template_ids[2].append(str(response_mobil1.json()['id']))
        return template_ids


    def __create_ads_cpc(self, company_ids, link_url_vorota, link_url_goal, link_url_mobil1, template_ids, token):
        ads_ids = []
        __header = {'Host': 'target.my.com', 'Content-Type': 'application/json',
                  'Accept-Encoding': 'gzip,deflate,compress',
                  'Authorization': 'Bearer 75kH6nlRjszveEBvR1jjH8667ykcluDt9BabULHPECGNwcyRKonvBPvBpAdkwD7tY2Ukocy0Qq6AdDxcKOTi6zrnjd9NwX3Igp8eh0eGw6AM1dopXtze6bk6Q22XyWzZ3H0XfulieOkRg0SPshLrNFdC60mIEd4P3C9VUoU2KgGrYBu3OKF1weUsEp8p9Bo4SQyQx1ofWMvTUzDv0q2IjczfB'}
        for i in range(len(company_ids)):
            __text_1 = {'campaign': {'id': company_ids[i]}, 'content': {
                'html5': {'id': template_ids[0][i]}},
                    'status': 'active',
                    'textblocks': {}, 'url': str(link_url_vorota[i]),
                    'system_status': 'active',
                    'banner_fields': ['content', 'textblocks', 'urls'],
                    'urls': {'link_1': {'url': str(link_url_vorota[i]), 'url_object_type': 'domain'}}}

            __text_2 = {'campaign': {'id': company_ids[i]}, 'content': {
                'html5': {'id': template_ids[1][i]}},
                      'status': 'active',
                      'textblocks': {}, 'url': str(link_url_goal[i]),
                      'system_status': 'active',
                      'banner_fields': ['content', 'textblocks', 'urls'],
                      'urls': {'link_1': {'url': str(link_url_goal[i]), 'url_object_type': 'domain'}}}

            __text_3 = {'campaign': {'id': company_ids[i]}, 'content': {
                'html5': {'id': template_ids[2][i]}},
                      'status': 'active',
                      'textblocks': {}, 'url': str(link_url_mobil1[i]),
                      'system_status': 'active',
                      'banner_fields': ['content', 'textblocks', 'urls'],
                      'urls': {'link_1': {'url': str(link_url_mobil1[i]), 'url_object_type': 'domain'}}}

            response = requests.post('https://target.my.com/api/v1/campaigns/' + str(company_ids[i]) +
                                     '/banners.json', data=json.dumps([__text_1, __text_2, __text_3]), headers=__header)
            print(response.json())
            ads_ids.append(response.json()[0]['id'])
        return ads_ids

    def __create_ads(self, company_ids, link_url, template_ids, token):
        ads_ids = []
        __header = {'Host': 'target.my.com', 'Content-Type': 'application/json',
                  'Accept-Encoding': 'gzip,deflate,compress',
                  'Authorization': 'Bearer 7Db55KfCPP0vpYzyWna8mmfOwGwSmizWGgDECvNzT4OMSNIL1c3BqyWuJKyC5sr8ahL2oPDbGoRvYhRlOHblrcnnVQSjSB463vc2xTTU3l9Z8HckazKcMWJcpQFWgHPLESG3t5TsbJ0m1IhK5c2toMT3Zc8LODINPlGPDDgeNYkgwMLtV9hQGfZWkwCOFL6LCPiofY9bGKtNeHMtNH94cxAa5uHHtomY'}
        print(template_ids)
        print(link_url)
        for i in range(len(company_ids)):
            __text = {'campaign': {'id': company_ids[i]}, 'content': {
                'html5': {'id': template_ids[i]}},
                    'status': 'active',
                    'textblocks': {}, 'url': str(link_url[i]),
                    'system_status': 'active',
                    'banner_fields': ['content', 'textblocks', 'urls'],
                    'urls': {'link_1': {'url': str(link_url[i]), 'url_object_type': 'domain'}}}
            response = requests.post('https://target.my.com/api/v1/campaigns/' + str(company_ids[i]) +
                                     '/banners.json', data=json.dumps(__text), headers=__header)
            print(response.json())
            ads_ids.append(response.json()['id'])
        return ads_ids

    def start(self):
        token = str(self.text_2.get())
        try:
            name, company_day_limit, company_all_limit, start_time, stop_time, \
            cpm, cpc, link_url, pixels, link_template, sex, age_from, age_to, geo = self.get_columns_data()

            #name, company_day_limit, company_all_limit, start_time, stop_time, \
            #cpm, cpc, link_template_vorota, link_url_vorota, \
            #link_template_goal, link_url_goal, link_template_mobil1, link_url_mobil1, pixels, \
            #sex, age_from, age_to, geo = self.get_columns_data_cpc()
        except FileNotFoundError:
            message = 'Укажите путь к таблице!'
            ctypes.windll.user32.MessageBoxW(0, message, 'Ads & Campaigns', 0)
            return 0
        address, radius, lng, lat = self.__get_geo(geo)

        __company_ids = self.__create_campaigns(name, company_day_limit, company_all_limit, start_time, stop_time, cpm,
                                                 cpc, sex, age_from, age_to, address, radius, lng, lat, pixels, token)

        print(len(__company_ids))
        print(__company_ids)
        #template_ids = self.__create_template_ids_cpc(link_template_vorota, link_template_goal, link_template_mobil1)
        #__ads_ids = self.__create_ads_cpc(__company_ids, link_url_vorota, link_url_goal, link_url_mobil1
                                      #, template_ids, token)

        template_ids = self.__create_template_ids(link_template)

        __ads_ids = self.__create_ads(__company_ids, link_url, template_ids, token)
        print(__ads_ids)

        __message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, __message, 'Ads & Campaigns', 0)
        print('ok')

        return {}

    def vk_token(self):
        header = {'Host': 'target.my.com', 'Content-Type': 'application/x-www-form-urlencoded'}
        '''
        Делает запрос к БД, смотрит, если там такие секрет и ид.
        Если есть, то берет их токен, делает пробный запрос на что-либо.
        Если работает - ОК
        Если нет - Идет ниже и рефрешит токен
        Если таких ид и секрет нет, то получает токен и вносит его вместе с данными в БД
        '''
        access_token_req = {
            "client_id": "n5zvIiexebHHazsF",
            "client_secret": "NoETjJTbsmAyE752NUakv2bpExZii88W2U1wddWPywayBcj2sjDVMr4fzFyvY87HxWtt7tJMoUNPY6K6aI2U3PjjROJRisTbzQMbfLiogsuN70631Bb3VoBxsIbUIroICpPCAJBfJ1aGijYianGKfzVuwLDn1esiTFntX5Gm0dWPA6XjYenpCXfrMrLbONcU8I2MEDTg5mGzPsQgpdMekjA5gLtHSfxJokhoN8",
            "refresh_token": "hzmqAFamDa9citPSbeVN9k4gRMTfezfsxZ1IhlPXBZ22aCISNENB80tix0pXLGjRsiRW6Fy1qf24NVALgg6CIthJ4sILY0gYoBXZaf40aR8TfKOqdm5yyuQXkRr21eBKDyWhAINauogMs8L2PtDJZNDSXaF3k8O2zrwgDznC2ho4wuPLe9SuNRkKNB9wpD079BK7rfZtO1x",
            "grant_type": "refresh_token"
        }

        r = requests.post('https://target.my.com/api/v2/oauth2/token.json', data=access_token_req, headers=header)

        print(r.json())

    def dialog(self, parent):
        self.parent = parent

        # Create main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_new_excel_file)

        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        self.label_2 = Label(parent, text="Укажите client id")
        self.text_2 = Entry(parent)

        self.label_2.pack()
        self.text_2.pack()

        self.label_3 = Label(parent, text="Укажите client secret")
        self.text_3 = Entry(parent)

        self.label_3.pack()
        self.text_3.pack()

        self.but_auth_ok = Button(parent, text="Получить токен", command=self.vk_token)
        self.but_auth_ok.pack()

        # Start button
        self.but_start = Button(parent, text="Создать кампанию/объявление", command=self.start)
        self.but_start.pack()


if __name__ == "__main__":
    root = Tk()
    root.minsize(width=500, height=400)
    gui = Gui(root)
    root.mainloop()
