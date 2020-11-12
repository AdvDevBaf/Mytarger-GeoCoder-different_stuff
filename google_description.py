import pandas as pd
import requests
import openpyxl
import time
from tkinter.filedialog import askopenfilename
from collections import defaultdict


path = str(askopenfilename())

try:
    wb = openpyxl.load_workbook(str(path))
except FileNotFoundError:
    message = 'Укажите путь к таблице!'
sheet = wb.worksheets[0]
address = []

for i in range(1, sheet.max_row):
    if sheet.cell(row=i, column=1).value is None:
        max_row = i - 1
        print(max_row)
        break
    else:
        max_row = sheet.max_row
        print(max_row)

for i in range(2, max_row + 1):
    address.append(str(sheet.cell(row=i, column=1).value))


print(address)
print('da')
description = []
desc = defaultdict(list)
print(desc)
for i in range(0,len(address)):
    try:
        print(address[i])
        response = requests.get("https://data.42matters.com/api/v2.0/android/apps/lookup.json?p="+str(address[i])+"&fields=similar&lang=en&access_token=1300b9852763b84f7ae8326465e62bbfba660c3c")
        print(response.json())
        desc[str(address[i])]=[]
        for j in range(0,len(response.json()['similar'])):
            print(response.json()['similar'][0])
            description.append(response.json()['similar'][j])
            desc[address[i]].append(str(response.json()['similar'][j]))
        print(desc)
    except KeyError:
        time.sleep(10)
        print('KeyError: token is expired or app was deleted')
        print(print(response.json()))

print(desc)

wb = openpyxl.load_workbook(str(path))
sheet = wb.worksheets[0]
for i in range(0, len(address)):
    sheet.cell(row=i + 2, column=5).value = str(desc[address[i]])
wb.save(str(path))
print('ok')

