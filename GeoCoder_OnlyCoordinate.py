import requests
import hashlib
import json
from requests_toolbelt.multipart.encoder import MultipartEncoder
from tkinter import *
from tkinter.filedialog import askopenfilename
import openpyxl
import ctypes
import time


class Yandex:
    def get_adress(self, address):
        yandex_address = []
        street = []
        latitude = []
        longitude = []
        coordinate = []
        print(address)
        time.sleep(12)
        for i in range(len(address)):
            response = requests.get("https://geocode-maps.yandex.ru/1.x/?apikey=55080292-108d-4b98-b077-fb1c2af3affd&format=json&geocode="+str(address[i])).json()
            print('tak')
            print(address[i])
            print(response)
            print(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country']['AddressLine'])

            print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']).split(" "))
            coordinate.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']).split(" "))
            print(coordinate)
            print(coordinate)
        print(coordinate)
        print(yandex_address)
        print(street)
        print(coordinate)
        for i in range(len(coordinate)):
            latitude.append(coordinate[i][1])
            longitude.append(coordinate[i][0])
        return yandex_address, street, latitude, longitude


class Gui(Toplevel, Yandex):
    def __init__(self, parent, title="Работа с API"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def search_folder_for_excel_file(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def get_columns_data(self):
        # Получим столбцы из эксель файла
        try:
            wb = openpyxl.load_workbook(str(self.text_1.get()))
        except FileNotFoundError:
            message = 'Укажите путь к таблице!'
            ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
        sheet = wb.worksheets[0]
        address = []

        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value is None:
                max_row = i - 1
                print(max_row)
                break
            else:
                max_row = sheet.max_row
                print(max_row)
        max_row = 200
        for i in range(2, max_row + 1):
            address.append(sheet.cell(row=i, column=1).value)

        return address

    def refill_table(self, latitude, longitude):
        wb = openpyxl.load_workbook(str(self.text_1.get()))
        sheet = wb.worksheets[0]
        for i in range(0, len(latitude)):
            print(latitude[i])
            sheet.cell(row=i+2, column=6).value = str(latitude[i])
            sheet.cell(row=i+2, column=7).value = str(longitude[i])
        wb.save(str(self.text_1.get()))
        message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, message, 'Работа с API', 0)
        print('ok')
        return {}

    def start(self):
        address = self.get_columns_data()
        yandex_address, street, latitude, longitude = self.get_adress(address)

        return self.refill_table(latitude, longitude)

    def dialog(self, parent):
        self.parent = parent

        # Created main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_excel_file)

        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        # start button
        self.but_start = Button(parent, text="Получить адреса", command=self.start)
        self.but_start.pack()


if __name__ == "__main__":
        root = Tk()
        root.minsize(width=500, height=400)
        gui = Gui(root)
        root.mainloop()
