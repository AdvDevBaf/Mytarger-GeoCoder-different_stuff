import pandas as pd
import requests
import openpyxl
import time
import urllib
from bs4 import BeautifulSoup

class AppURLopener(urllib.request.FancyURLopener):
    version = "Mozilla/5.0"


for z in range(10,15):
    path = str('C:\\Users\\AMasanov\\Desktop\\Игры по возрастам - Copy\\games for ' + str(z) + ' years old girls.xlsx')

    wb = openpyxl.load_workbook(str(path))
    sheet = wb.worksheets[0]
    address = []

    for j in range(2, sheet.max_row+1):
        if sheet.cell(row=j, column=5).value is None:
            max_row = j - 1
            break
        else:
            max_row = sheet.max_row

    print(max_row)

    for k in range(2, max_row + 1):
        address.append(sheet.cell(row=k, column=5).value)

    adv = []
    links = []
    for l in range(len(address)):
        links.append("https://play.google.com/store/apps/details?id=" + str(address[l]))

    names = []

    for link in links:
        opener = AppURLopener()
        print(link)
        sas = requests.get(link).text
        soup = BeautifulSoup(sas)
        soup_list = soup.findAll('title')
        print(soup_list[0])
        print('da')
        names.append(str(soup_list[0]).replace('<title id="main-title">','').replace('</title>',''))

    for m in range(0, len(address)):
        try:
            page = requests.get('https://play.google.com/store/apps/details?id=' + str(address[m]))
            if 'Contains Ads' in page.text:
                print('yes')
                adv.append('yes')
            else:
                print('no')
                adv.append('no')
        except:
            adv.append('no')
            print('Something happen')
            time.sleep(10)

    '''
    описание добавить еще
    '''


    table = pd.DataFrame({'id': address, 'Название': names, 'Ссылка': links, 'Реклама в приложении': adv})
    table.to_csv(str("C:\\Users\AMasanov\\Desktop\\Игры по возрастам - Расширенный") + '/' + str("games for "+ str(z) +" years old girls") + '.csv', sep=';',
                 index=False,
                 encoding='utf-8-sig')
    print('ok')