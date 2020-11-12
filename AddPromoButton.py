import json
from tkinter import *
from tkinter.filedialog import askopenfilename
import requests
import openpyxl
import ctypes
import time


class Gui(Toplevel):
    def __init__(self, parent, title="Create ads & campaigns"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def search_folder_for_new_excel_file(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def get_columns_data_cpc(self):
        from datetime import datetime
        # Get cells from excel file
        wb = openpyxl.load_workbook(str(self.text_1.get()))
        sheet = wb.worksheets[0]

        # Create lists for campaigns
        name = []
        company_day_limit = []
        company_all_limit = []
        start_time = []
        stop_time = []

        cpm = []
        description = []
        title = []
        text = []
        sex = []
        age_from = []
        age_to = []
        geo = []
        link_template_600 = []
        link_template_256 = []
        link_template_1080 = []
        link_url = []

        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value is None:
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(2, max_row + 1):
            # Fill campaign lists
            name.append(sheet.cell(row=i, column=1).value)
            sex.append(sheet.cell(row=i, column=2).value)
            age_from.append(sheet.cell(row=i, column=3).value)
            age_to.append(sheet.cell(row=i, column=4).value)
            geo.append(str(sheet.cell(row=i, column=5).value))
            company_day_limit.append(str(sheet.cell(row=i, column=8).value))
            company_all_limit.append(str(sheet.cell(row=i, column=9).value))
            cpm.append(str(sheet.cell(row=i, column=10).value).replace(',', '.'))
            description.append(str(sheet.cell(row=i, column=11).value).replace(',', '.'))
            title.append(str(sheet.cell(row=i, column=12).value).replace(',', '.'))
            text.append(str(sheet.cell(row=i, column=13).value).replace(',', '.'))
            link_template_600.append(str(sheet.cell(row=i, column=14).value).replace('\\', '/'))
            link_template_256.append(str(sheet.cell(row=i, column=15).value).replace('\\', '/'))
            link_template_1080.append(str(sheet.cell(row=i, column=16).value).replace('\\', '/'))
            link_url.append(str(sheet.cell(row=i, column=17).value).replace('\\', '/'))
            #link_template_mobil1.append(str(sheet.cell(row=i, column=16).value).replace('\\', '/'))
            #link_url_mobil1.append(str(sheet.cell(row=i, column=17).value).replace('\\', '/'))

            start = datetime.strptime(str(sheet.cell(row=i, column=6).value)[:10].replace('-', '.'), '%Y.%m.%d')
            stop = datetime.strptime(str(sheet.cell(row=i, column=7).value)[:10].replace('-', '.'), '%Y.%m.%d')
            start_time.append(str(start.strftime('%d.%m.%Y')))
            stop_time.append(str(stop.strftime('%d.%m.%Y')))


        print(start_time)
        print(stop_time)
        print(cpm)
        return name, company_day_limit, company_all_limit, start_time, stop_time,\
               cpm, description, title, text, link_template_600, link_template_256, \
               link_template_1080, link_url, \
               sex, age_from, age_to, geo

    def __refresh_token(self):
        pass  # Получить/Обновить токен

    def __get_geo(self, geo):
        address = []
        radius = []
        lng = []
        lat = []
        for element in geo:
            location = element.split(':')
            location = location[::-1]
            address.append(str(location[2]))
            radius.append(str(location[1])[:len(location[1])-1])
            coord = str(location[0]).split(',')
            lng.append(coord[1])
            lat.append(coord[0])
        print(address)
        print(radius)
        print(lng)
        print(lat)
        return address, radius, lng, lat

    def __create_campaigns(self, name, company_day_limit, company_all_limit, start_time, stop_time, cpm,
                           sex, age_from, age_to, address, radius, lng, lat, token):
        print(age_from[0])
        print(age_to[0])
        count = 0
        campaign_ids = []
        campaign_names = []
        __header = {'Host': 'target.my.com', 'Content-Type': 'application/json',
                    'Accept-Encoding': 'gzip,deflate,compress',
                    'Authorization': 'Bearer DIFGlOUSyrLAPVu8aOi6SfOnFkUTPdUDxeDVjI6mEbBz0Osonlb1jPl49INFX9su6yQmyjdy93zjRfn7MnEweWmHcp3gmtW3iepUVX6rlaMhbKiHZj4pzCZju7tcySbjL4Fz6hATBv9KkVS5ZjYJPpvnqGdl3RFgftsgvRbTUfmiEpGLxWLcvfiGhMySUB6I0V1YrbvBrr0RHkAjgN5QiJ5BKQ8i0fLo4VZq6inSiM29hhUld9OKeReaT7Br5F'}
        __sex = {
            "Мужской": "M",
            "Женский": "F",
            "Любой": "MF"
        }

        # 503 20290 20292 | 504 20291 20289
        # desktop_site_240x400_cpc_ttm_mailru
        for i in range(len(name)):
            if str(name[i]) not in campaign_names:
                campaign_names.append(str(name[i]))
                print('net')
                company = {"name": str(name[i]), "price": str(cpm[i]), "budget_limit_day": str(company_day_limit[i]),
                           "budget_limit": str(company_all_limit[i]), "date_start": str(start_time[i]),
                           "date_end": str(stop_time[i]), "package": {"id": 646},
                           "targetings": {"regions": [188], "sex": __sex[sex[i]],
                                          "age": [x for x in range(age_from[i], age_to[i]+1)],
                                          "pads": [{"id": 76920}, {'id': 76924}],
                                          "local_geo": {"type": "now", "regions": [{"lng": str(lng[i]),  # долгота
                                                                                        "lat": str(lat[i]),  # широта
                                                                                        "radius": str(radius[i]),
                                                                                        "label": str(address[i])}]
                                                            }
                                              }
                               }

                response = requests.post('https://target.my.com/api/v1/campaigns.json', data=json.dumps(company),
                                         headers=__header)
                count += 1
                print(count)
                if count == 179:
                    import datetime
                    now = datetime.datetime.now()
                    a = (60 - now.minute)*60
                    print(a)
                    time.sleep(a)
                print(response.json())
                print(response.json()['id'])
                campaign_ids.append(response.json()['id'])
                print(campaign_ids)
            else:
                campaign_ids.append(campaign_names.index(str(name[i])))

        return campaign_ids

    def __create_template_ids_urls(self, link_template_600, link_template_256, link_template_1080):
        print(len(link_template_256))
        print(len(link_template_600))
        print(len(link_template_1080))
        template_ids = [[], [], []]
        template_urls = [[], [], []]
        headers = {
            'Host': 'target.my.com',
            'Authorization': 'Bearer DIFGlOUSyrLAPVu8aOi6SfOnFkUTPdUDxeDVjI6mEbBz0Osonlb1jPl49INFX9su6yQmyjdy93zjRfn7MnEweWmHcp3gmtW3iepUVX6rlaMhbKiHZj4pzCZju7tcySbjL4Fz6hATBv9KkVS5ZjYJPpvnqGdl3RFgftsgvRbTUfmiEpGLxWLcvfiGhMySUB6I0V1YrbvBrr0RHkAjgN5QiJ5BKQ8i0fLo4VZq6inSiM29hhUld9OKeReaT7Br5F',
            'Accept-Encoding': 'gzip,deflate,compress',
        }
        for i in range(len(link_template_600)):
            file_600 = {
                'file': (str(link_template_600[i]),
                         open(str(link_template_600[i]), 'rb')),
                'data': (None, '{"width": 600, "height": 600}'),
            }

            response_600 = requests.post('https://target.my.com/api/v2/content/static.json', files=file_600,
                                         headers=headers)

            file_256 = {
                'file': (str(link_template_256[i]),
                         open(str(link_template_256[i]), 'rb')),
                'data': (None, '{"width": 256, "height": 256}'),
            }

            response_256 = requests.post('https://target.my.com/api/v2/content/static.json', files=file_256,
                                         headers=headers)

            file_1080 = {
                'file': (str(link_template_1080[i]),
                         open(str(link_template_1080[i]), 'rb')),
                'data': (None, '{"width": 1080, "height": 607}'),
            }

            response_1080 = requests.post('https://target.my.com/api/v2/content/static.json', files=file_1080,
                                         headers=headers)

            print(response_600.json())
            print(response_256.json())
            print(response_1080.json())
            template_ids[0].append(str(response_600.json()['id']))
            template_ids[1].append(str(response_256.json()['id']))
            template_ids[2].append(str(response_1080.json()['id']))
            print(template_ids)
            template_urls[0].append(str(response_600.json()['variants']['original']['url']))
            template_urls[1].append(str(response_256.json()['variants']['original']['url']))
            template_urls[2].append(str(response_1080.json()['variants']['original']['url']))
            print(template_urls)
        return template_ids, template_urls

    def __create_urls(self, template_urls):
        print(len(template_urls))
        urls = [[], [], []]
        print('da')
        print(template_urls)

        headers = {
            'Host': 'target.my.com',
            'Authorization': 'Bearer DIFGlOUSyrLAPVu8aOi6SfOnFkUTPdUDxeDVjI6mEbBz0Osonlb1jPl49INFX9su6yQmyjdy93zjRfn7MnEweWmHcp3gmtW3iepUVX6rlaMhbKiHZj4pzCZju7tcySbjL4Fz6hATBv9KkVS5ZjYJPpvnqGdl3RFgftsgvRbTUfmiEpGLxWLcvfiGhMySUB6I0V1YrbvBrr0RHkAjgN5QiJ5BKQ8i0fLo4VZq6inSiM29hhUld9OKeReaT7Br5F',
            'Accept-Encoding': 'gzip,deflate,compress',
        }
        for i in range(len(template_urls[0])):
            response_600 = requests.post('https://target.my.com/api/v2/urls.json',
                                         data=json.dumps({'url': str(template_urls[0][i])}), headers=headers)
            print(response_600.json())
            urls[0].append(response_600.json()['id'])
            print(urls)

            response_256 = requests.post('https://target.my.com/api/v2/urls.json',
                                         data=json.dumps({'url': str(template_urls[1][i])}), headers=headers)
            print(response_256.json())
            urls[1].append(response_256.json()['id'])
            print(urls)

            response_1080 = requests.post('https://target.my.com/api/v2/urls.json',
                                         data=json.dumps({'url': str(template_urls[2][i])}), headers=headers)
            print(response_1080.json())
            urls[2].append(response_1080.json()['id'])
            print(urls)
        return urls

    def __create_ads(self, company_ids, link_url, template_ids, urls_ids, description, title, text, token):
        ads_ids = []
        __header = {'Host': 'target.my.com', 'Content-Type': 'application/json',
                  'Accept-Encoding': 'gzip,deflate,compress',
                  'Authorization': 'Bearer DIFGlOUSyrLAPVu8aOi6SfOnFkUTPdUDxeDVjI6mEbBz0Osonlb1jPl49INFX9su6yQmyjdy93zjRfn7MnEweWmHcp3gmtW3iepUVX6rlaMhbKiHZj4pzCZju7tcySbjL4Fz6hATBv9KkVS5ZjYJPpvnqGdl3RFgftsgvRbTUfmiEpGLxWLcvfiGhMySUB6I0V1YrbvBrr0RHkAjgN5QiJ5BKQ8i0fLo4VZq6inSiM29hhUld9OKeReaT7Br5F'}
        print(template_ids)
        print(link_url)
        for i in range(len(company_ids)):
            __text = {'content': {'icon_square': {'id': template_ids[0][i]}, 'image': {'id': template_ids[1][i]},
                                'promo_image': {'id': template_ids[2][i]}},
                    'textblocks': {
                        'primary': {'text': str(text[i]),
                                    'title': str(title[i])},
                    'about_company': {'text': '',
                                      'title': ''}},
                    'urls': {'icon_square': {'id': int(urls_ids[0][i])},
                             'image': {'id': int(urls_ids[1][i])},
                             'promo_image': {'id': int(urls_ids[2][i])},
                             'primary': {'url': str(link_url[i])}},
                      'call_to_action': 'learnMore'}
            response = requests.post('https://target.my.com/api/v1/campaigns/' + str(company_ids[i]) +
                                     '/banners.json', data=json.dumps(__text), headers=__header)
            print(response.json())
            ads_ids.append(response.json()['id'])
            print(ads_ids)
        return ads_ids

    def start(self):
        token = str(self.text_2.get())
        try:

            name, company_day_limit, company_all_limit, start_time, stop_time, \
            cpm, description, title, text, link_template_600, link_template_256, \
            link_template_1080, link_url, \
            sex, age_from, age_to, geo = self.get_columns_data_cpc()
        except FileNotFoundError:
            message = 'Укажите путь к таблице!'
            ctypes.windll.user32.MessageBoxW(0, message, 'Ads & Campaigns', 0)
            return 0
        address, radius, lng, lat = self.__get_geo(geo)

        #__company_ids = self.__create_campaigns(name, company_day_limit, company_all_limit, start_time, stop_time, cpm,
         #                                       sex, age_from, age_to, address, radius, lng, lat, token)
        __company_ids = [13087253, 13087254, 13087255, 13087256, 13087257, 13087258, 13087259, 13087260, 13087262, 13087271, 13087273, 13087274, 13087275, 13087276, 13087277, 13087278, 13087279, 13087280, 13087281, 13087282, 13087283, 13087284, 13087285, 13087286, 13087287, 13087288, 13087289, 13087291, 13087292, 13087293, 13087294, 13087295, 13087297, 13087298, 13087299, 13087301, 13087302, 13087303, 13087304, 13087305, 13087306, 13087307, 13087308, 13087310, 13087311, 13087312, 13087313, 13087314, 13087315, 13087316, 13087317, 13087318, 13087319, 13087320, 13087322, 13087323, 13087324, 13087325, 13087326]


        #print(len(__company_ids))
        #print(__company_ids)
        template_ids, template_urls = self.__create_template_ids_urls(link_template_600, link_template_256,
                                                                      link_template_1080)
        #print(template_ids)
        #print(template_urls)
        #template_idss = [['3276850', '3276853', '3276856', '3276858', '3276860', '3276864', '3276870', '3276873', '3276877', '3276880', '3275927', '3277297', '3277299', '3277302', '3277305', '3277307', '3277309', '3277312', '3277315', '3277318', '3277321', '3277324', '3277326', '3277329', '3277332', '3277337', '3277341', '3277345', '3277348', '3277351', '3277354', '3277356', '3277359', '3277362', '3277365', '3277368', '3277371', '3277374', '3277379', '3277381', '3277383', '3277385', '3277387', '3277390', '3277392', '3277395', '3277399', '3277401', '3277405', '3277408', '3277411', '3277414', '3277416', '3277418', '3277422', '3277424', '3277426', '3277428', '3277432']]

        #template_ids0 = list(map(int, template_idss[0]))
        #template_ids1 = list(map(int, template_idss[1]))
        #template_ids2 = list(map(int, template_idss[2]))
        #template_ids = [template_ids0,template_ids1,template_ids2]
        #template_urls = [['https://r.mradx.net/img/F3/DF43C8.png', 'https://r.mradx.net/img/94/8E167C.png', 'https://r.mradx.net/img/64/2EC78F.png', 'https://r.mradx.net/img/82/002A75.png', 'https://r.mradx.net/img/41/324FA3.png', 'https://r.mradx.net/img/22/25C25D.png', 'https://r.mradx.net/img/88/6B3B7B.png', 'https://r.mradx.net/img/59/7D4948.png', 'https://r.mradx.net/img/5B/893457.png', 'https://r.mradx.net/img/63/57B7AC.png', 'https://r.mradx.net/img/D1/F5E7A5.png', 'https://r.mradx.net/img/09/F3D69E.png', 'https://r.mradx.net/img/93/81BC03.png', 'https://r.mradx.net/img/9E/CC510B.png', 'https://r.mradx.net/img/BF/3349D0.png', 'https://r.mradx.net/img/82/362354.png', 'https://r.mradx.net/img/B7/C41F7B.png', 'https://r.mradx.net/img/82/320383.png', 'https://r.mradx.net/img/59/C853FF.png', 'https://r.mradx.net/img/E3/44DB4D.png', 'https://r.mradx.net/img/D8/FD194E.png', 'https://r.mradx.net/img/71/64685D.png', 'https://r.mradx.net/img/0E/F672C9.png', 'https://r.mradx.net/img/D3/12863B.png', 'https://r.mradx.net/img/E7/AA8701.png', 'https://r.mradx.net/img/C0/89C180.png', 'https://r.mradx.net/img/D9/936A11.png', 'https://r.mradx.net/img/DE/B2ECCB.png', 'https://r.mradx.net/img/C9/C5E057.png', 'https://r.mradx.net/img/F7/278730.png', 'https://r.mradx.net/img/9A/E9581A.png', 'https://r.mradx.net/img/A2/AF084F.png', 'https://r.mradx.net/img/6A/C757C0.png', 'https://r.mradx.net/img/7A/3B6786.png', 'https://r.mradx.net/img/D1/4057B3.png', 'https://r.mradx.net/img/41/4D3D37.png', 'https://r.mradx.net/img/E2/F27620.png', 'https://r.mradx.net/img/86/D9F0DB.png', 'https://r.mradx.net/img/A2/059F2C.png', 'https://r.mradx.net/img/DC/B6FBA9.png', 'https://r.mradx.net/img/28/ABBB33.png', 'https://r.mradx.net/img/ED/03AB03.png', 'https://r.mradx.net/img/20/0CC2EA.png', 'https://r.mradx.net/img/BF/1CAFB6.png', 'https://r.mradx.net/img/88/358115.png', 'https://r.mradx.net/img/C3/E8592F.png', 'https://r.mradx.net/img/F1/647F37.png', 'https://r.mradx.net/img/38/615BDD.png', 'https://r.mradx.net/img/37/674DCA.png', 'https://r.mradx.net/img/1D/168282.png', 'https://r.mradx.net/img/58/94F17E.png', 'https://r.mradx.net/img/FA/4314B1.png', 'https://r.mradx.net/img/76/DFDAAE.png', 'https://r.mradx.net/img/C9/91EA20.png', 'https://r.mradx.net/img/D2/4BA415.png', 'https://r.mradx.net/img/79/12C445.png', 'https://r.mradx.net/img/78/DADD7D.png', 'https://r.mradx.net/img/DE/23C9F2.png', 'https://r.mradx.net/img/F3/A2DB18.png'], ['https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png', 'https://r.mradx.net/img/F2/23E27F.png'], ['https://r.mradx.net/img/74/F40CBF.png', 'https://r.mradx.net/img/23/F19601.png', 'https://r.mradx.net/img/74/252C4F.png', 'https://r.mradx.net/img/E5/AB36F6.png', 'https://r.mradx.net/img/C8/AB559B.png', 'https://r.mradx.net/img/06/75D167.png', 'https://r.mradx.net/img/2B/9B26F4.png', 'https://r.mradx.net/img/E6/3FD071.png', 'https://r.mradx.net/img/AB/FEA213.png', 'https://r.mradx.net/img/AD/BE93CD.png', 'https://r.mradx.net/img/37/00D247.png', 'https://r.mradx.net/img/82/66F178.png', 'https://r.mradx.net/img/F5/98B207.png', 'https://r.mradx.net/img/E2/12D109.png', 'https://r.mradx.net/img/AE/2F53BA.png', 'https://r.mradx.net/img/F1/C657CE.png', 'https://r.mradx.net/img/F3/9ED0C8.png', 'https://r.mradx.net/img/BA/CE1BAA.png', 'https://r.mradx.net/img/A5/505FD2.png', 'https://r.mradx.net/img/8F/34154F.png', 'https://r.mradx.net/img/3B/CA4EB9.png', 'https://r.mradx.net/img/DA/FA5716.png', 'https://r.mradx.net/img/2F/BD45E1.png', 'https://r.mradx.net/img/14/492209.png', 'https://r.mradx.net/img/8A/EDBBEA.png', 'https://r.mradx.net/img/3D/7F1449.png', 'https://r.mradx.net/img/B7/09DC5B.png', 'https://r.mradx.net/img/D9/0CC9A9.png', 'https://r.mradx.net/img/B5/839805.png', 'https://r.mradx.net/img/7E/DCB8BF.png', 'https://r.mradx.net/img/DE/4DD282.png', 'https://r.mradx.net/img/59/8A7FA6.png', 'https://r.mradx.net/img/17/43914B.png', 'https://r.mradx.net/img/F6/69B362.png', 'https://r.mradx.net/img/7D/0CB3FF.png', 'https://r.mradx.net/img/CC/4B7551.png', 'https://r.mradx.net/img/07/D5AE9F.png', 'https://r.mradx.net/img/30/3D4896.png', 'https://r.mradx.net/img/DC/1648FC.png', 'https://r.mradx.net/img/80/D7B755.png', 'https://r.mradx.net/img/7C/C58874.png', 'https://r.mradx.net/img/5A/427E3E.png', 'https://r.mradx.net/img/32/F10593.png', 'https://r.mradx.net/img/D9/080F69.png', 'https://r.mradx.net/img/55/95A425.png', 'https://r.mradx.net/img/1A/DD57EE.png', 'https://r.mradx.net/img/3D/A7CF76.png', 'https://r.mradx.net/img/B2/14A9FA.png', 'https://r.mradx.net/img/BD/E72A4B.png', 'https://r.mradx.net/img/24/51ABEA.png', 'https://r.mradx.net/img/58/6D324F.png', 'https://r.mradx.net/img/53/34CF72.png', 'https://r.mradx.net/img/EE/A7B448.png', 'https://r.mradx.net/img/38/DA182F.png', 'https://r.mradx.net/img/35/A3823D.png', 'https://r.mradx.net/img/31/93A4C4.png', 'https://r.mradx.net/img/94/EE9607.png', 'https://r.mradx.net/img/EB/675239.png', 'https://r.mradx.net/img/90/A16E9A.png']]
        urls_ids = self.__create_urls(template_urls)
        print(template_ids)
        print(template_urls)
        print(urls_ids)
        #urls_ids = [[15158029, 15158032, 15158034, 15158036, 15158038, 15158040, 15158042, 15158044, 15158046, 15158048, 15158050, 15158052, 15158054, 15158056, 15158058, 15158060, 15158062, 15158064, 15158066, 15158068, 15158070, 15158072, 15158074, 15158076, 15158078, 15158080, 15158082, 15158085, 15158087, 15158089, 15158091, 15158093, 15158095, 15158097, 15158099, 15158101, 15158103, 15158105, 15158107, 15158110, 15158112, 15158114, 15158116, 15158118, 15158120, 15158122, 15158124, 15158126, 15158128, 15158130, 15158132, 15158134, 15158136, 15158138, 15158140, 15158142, 15158144, 15158146, 15158148, 15158150, 15158152, 15158155, 15158157, 15158159, 15158161, 15158163, 15158165, 15158167, 15158169, 15158171, 15158173, 15158175, 15158178, 15158180, 15158182, 15158184, 15158186, 15158188, 15158190, 15158192, 15158194, 15158196, 15158198, 15158200, 15158202, 15158204, 15158206, 15158208, 15158210, 15158212, 15158214, 15158216, 15158218, 15158220, 15158222, 15158224, 15158226, 15158228, 15158230, 15158232, 15158234, 15158236, 15158238, 15158240, 15158242, 15158244, 15158246, 15158248, 15158250, 15158252, 15158254, 15158256, 15158258, 15158260, 15158262, 15158264, 15158266, 15158268, 15158270, 15158272, 15158274, 15158276, 15158278, 15158280, 15158282, 15158284, 15158286, 15158288, 15158290, 15158292, 15158294, 15158296, 15158298, 15158300, 15158302, 15158304, 15158306, 15158308, 15158310, 15158312, 15158314, 15158316, 15158318, 15158320, 15158322, 15158324, 15158326, 15158328, 15158330, 15158332, 15158334, 15158336, 15158338, 15158340, 15158342, 15158344, 15158346, 15158348, 15158350, 15158352, 15158354, 15158356, 15158358, 15158360, 15158362, 15158429, 15158431, 15158433, 15158435, 15158438, 15158440, 15158443, 15158446, 15158448, 15158450, 15158452, 15158454, 15158456, 15158458, 15158101, 15158460, 15158462, 15158464, 15158466, 15158468, 15158470, 15158472, 15158474, 15158476, 15158478, 15158481, 15158483, 15158485, 15158487, 15158489, 15158491, 15158493, 15158495, 15158497, 15158499, 15158501, 15158503, 15158505, 15158507, 15158509, 15158511, 15158513, 15158515, 15158517, 15158519, 15158521, 15158523, 15158525, 15158527, 15158530, 15158532, 15158534, 15158536, 15158538, 15158540, 15158542, 15158544, 15158546, 15158548, 15158551, 15158553, 15158555, 15158557], [15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030, 15158030], [15158031, 15158033, 15158035, 15158037, 15158039, 15158041, 15158043, 15158045, 15158047, 15158049, 15158051, 15158053, 15158055, 15158057, 15158059, 15158061, 15158063, 15158065, 15158067, 15158069, 15158071, 15158073, 15158075, 15158077, 15158079, 15158081, 15158084, 15158086, 15158088, 15158090, 15158092, 15158094, 15158096, 15158098, 15158100, 15158102, 15158104, 15158106, 15158108, 15158111, 15158113, 15158115, 15158117, 15158119, 15158121, 15158123, 15158125, 15158127, 15158129, 15158131, 15158133, 15158135, 15158137, 15158139, 15158141, 15158143, 15158145, 15158147, 15158149, 15158151, 15158154, 15158156, 15158158, 15158160, 15158162, 15158164, 15158166, 15158168, 15158170, 15158172, 15158174, 15158176, 15158179, 15158181, 15158183, 15158185, 15158187, 15158189, 15158191, 15158193, 15158195, 15158197, 15158199, 15158201, 15158203, 15158205, 15158207, 15158209, 15158211, 15158213, 15158215, 15158217, 15158219, 15158221, 15158223, 15158225, 15158227, 15158229, 15158231, 15158233, 15158235, 15158237, 15158239, 15158241, 15158243, 15158245, 15158247, 15158249, 15158251, 15158253, 15158255, 15158257, 15158259, 15158261, 15158263, 15158265, 15158267, 15158269, 15158271, 15158273, 15158275, 15158277, 15158279, 15158281, 15158283, 15158285, 15158287, 15158289, 15158291, 15158293, 15158295, 15158297, 15158299, 15158301, 15158303, 15158305, 15158307, 15158309, 15158311, 15158313, 15158315, 15158317, 15158319, 15158321, 15158323, 15158325, 15158327, 15158329, 15158331, 15158333, 15158335, 15158337, 15158339, 15158341, 15158343, 15158345, 15158347, 15158349, 15158351, 15158353, 15158355, 15158357, 15158359, 15158361, 15158363, 15158430, 15158432, 15158434, 15158436, 15158439, 15158442, 15158445, 15158447, 15158449, 15158451, 15158453, 15158455, 15158457, 15158459, 15158102, 15158461, 15158463, 15158465, 15158467, 15158469, 15158471, 15158473, 15158475, 15158477, 15158480, 15158482, 15158484, 15158486, 15158488, 15158490, 15158492, 15158494, 15158496, 15158498, 15158500, 15158502, 15158504, 15158506, 15158508, 15158510, 15158512, 15158514, 15158516, 15158518, 15158520, 15158522, 15158524, 15158526, 15158529, 15158531, 15158533, 15158535, 15158537, 15158539, 15158541, 15158543, 15158545, 15158547, 15158550, 15158552, 15158554, 15158556, 15158558]]
        __ads_ids = self.__create_ads(__company_ids, link_url, template_ids, urls_ids,description, title,
                                      text, token)
        # [31744814, 31744816, 31744817, 31744818, 31744819, 31744820, 31744821, 31744822, 31744823, 31744825, 31744826, 31744827, 31744828, 31744829, 31744830, 31744832, 31744833, 31744834, 31744835, 31744836, 31744837, 31744838, 31744839, 31744843, 31744844, 31744845, 31744846, 31744847, 31744848, 31744849, 31744850, 31744851, 31744852, 31744853, 31744854, 31744855, 31744857, 31744858, 31744860, 31744861, 31744863, 31744864, 31744865, 31744866, 31744867, 31744869, 31744870, 31744871, 31744872, 31744873, 31744874, 31744875, 31744876, 31744877, 31744879, 31744881, 31744885, 31744886, 31744887, 31744888, 31744889, 31744890, 31744891, 31744892, 31744893, 31744895, 31744896, 31744897, 31744898, 31744899, 31744900, 31744901, 31744903, 31744904, 31744908, 31744916, 31744918, 31744919, 31744920, 31744921, 31744922, 31744923, 31744924, 31744925, 31744928, 31744929, 31744930, 31744931, 31744932, 31744933, 31744934, 31744935, 31744974, 31744977, 31744979, 31744980, 31744981, 31744983, 31744984, 31744985, 31744987, 31744988, 31744989, 31744990, 31744991, 31744992, 31744993, 31744994, 31744995, 31744996, 31744997, 31744998, 31745000, 31745001, 31745002, 31745003, 31745004, 31745005, 31745006, 31745008, 31745009, 31745010, 31745011, 31745014, 31745018, 31745019, 31745020, 31745021, 31745022, 31745023, 31745024, 31745025, 31745032, 31745033, 31745034, 31745035, 31745036, 31745037, 31745038, 31745040, 31745041, 31745042, 31745043, 31745045, 31745046, 31745047, 31745048, 31745049, 31745051, 31745052, 31745053, 31745059, 31745061, 31745062, 31745063, 31745065, 31745066, 31745068, 31745069, 31745083, 31745085, 31745089, 31745090, 31745092, 31745098, 31745099, 31745100, 31745101, 31745102, 31745103, 31745104, 31745105, 31745106, 31745107, 31745108, 31745109, 31745110, 31745111, 31745112, 31745113, 31745114, 31745115, 31745117, 31745118, 31745119, 31745120, 31745121, 31745122, 31745124, 31745125, 31745126, 31745127, 31745128, 31745129, 31745130, 31745131, 31745132, 31745133, 31745134, 31745135, 31745136, 31745137, 31745138, 31745139, 31745140, 31745141, 31745144, 31745149, 31745153, 31745154, 31745155, 31745156, 31745157, 31745158, 31745159, 31745161, 31745162, 31745163, 31745164, 31745166, 31745167, 31745174, 31745175, 31745176, 31745178, 31745179, 31745183, 31745184, 31745185, 31745187, 31745188, 31745189, 31745190, 31745191, 31745192, 31745194, 31745196, 31745197, 31745199, 31745200, 31745201, 31745203, 31745204, 31745208, 31745210, 31745212, 31745213, 31745214, 31745215, 31745216, 31745218, 31745219, 31745220, 31745221, 31745222, 31745223, 31745224, 31745226, 31745227, 31745228, 31745229, 31745230, 31745233, 31745234, 31745235, 31745237]
        #print(__ads_ids)

        __message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, __message, 'Ads & Campaigns', 0)
        print('ok')

        return {}

    def vk_token(self):
        header = {'Host': 'target.my.com', 'Content-Type': 'application/x-www-form-urlencoded'}
        '''
        Делает запрос к БД, смотрит, если там такие секрет и ид.
        Если есть, то берет их токен, делает пробный запрос на что-либо.
        Если работает - ОК
        Если нет - Идет ниже и рефрешит токен
        Если таких ид и секрет нет, то получает токен и вносит его вместе с данными в БД
        '''
        access_token_req = {
            "client_id": "n5zvIiexebHHazsF",
            "client_secret": "NoETjJTbsmAyE752NUakv2bpExZii88W2U1wddWPywayBcj2sjDVMr4fzFyvY87HxWtt7tJMoUNPY6K6aI2U3PjjROJRisTbzQMbfLiogsuN70631Bb3VoBxsIbUIroICpPCAJBfJ1aGijYianGKfzVuwLDn1esiTFntX5Gm0dWPA6XjYenpCXfrMrLbONcU8I2MEDTg5mGzPsQgpdMekjA5gLtHSfxJokhoN8",
            "refresh_token": "hzmqAFamDa9citPSbeVN9k4gRMTfezfsxZ1IhlPXBZ22aCISNENB80tix0pXLGjRsiRW6Fy1qf24NVALgg6CIthJ4sILY0gYoBXZaf40aR8TfKOqdm5yyuQXkRr21eBKDyWhAINauogMs8L2PtDJZNDSXaF3k8O2zrwgDznC2ho4wuPLe9SuNRkKNB9wpD079BK7rfZtO1x",
            "grant_type": "refresh_token"
        }

        r = requests.post('https://target.my.com/api/v2/oauth2/token.json', data=access_token_req, headers=header)

        print(r.json())

    def dialog(self, parent):
        self.parent = parent

        # Create main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_new_excel_file)

        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        self.label_2 = Label(parent, text="Укажите client id")
        self.text_2 = Entry(parent)

        self.label_2.pack()
        self.text_2.pack()

        self.label_3 = Label(parent, text="Укажите client secret")
        self.text_3 = Entry(parent)

        self.label_3.pack()
        self.text_3.pack()

        self.but_auth_ok = Button(parent, text="Получить токен", command=self.vk_token)
        self.but_auth_ok.pack()

        # Start button
        self.but_start = Button(parent, text="Создать кампанию/объявление", command=self.start)
        self.but_start.pack()


if __name__ == "__main__":
    root = Tk()
    root.minsize(width=500, height=400)
    gui = Gui(root)
    root.mainloop()
